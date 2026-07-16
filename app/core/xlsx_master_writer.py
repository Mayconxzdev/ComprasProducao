from __future__ import annotations

import os
import time
import json
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from .config import AppConfig
from .xlsx_loader_pro import detect_schema
from .utils_text import normalize_text
from .file_lock import cross_process_file_lock

_MASTER_SYNC_VERSION_FILE = "master_sync_version.json"



def _file_lock(lock_path: Path, timeout_sec: int = 10):
    return cross_process_file_lock(lock_path, timeout_sec=timeout_sec)


def _xlsx_master_path(config: AppConfig) -> Optional[Path]:
    p = (config.nas_master_path or "").strip()
    if p.lower().endswith((".xlsx", ".xlsm")):
        return Path(p)
    for source in config.xlsx_sources or []:
        value = str(source or "").strip()
        if value.lower().endswith((".xlsx", ".xlsm")):
            return Path(value)
    return None


def _write_master_sync_marker(xlsx: Path) -> None:
    marker = xlsx.parent / _MASTER_SYNC_VERSION_FILE
    payload = {
        "updated_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "updated_ts": int(time.time()),
        "master_file": xlsx.name,
    }
    tmp = marker.with_suffix(".tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(str(tmp), str(marker))


def _next_numeric_id(ws, id_col: int) -> str:
    max_id = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or id_col >= len(row):
            continue
        v = str(row[id_col] or "").strip()
        if v.isdigit():
            max_id = max(max_id, int(v))
    return str(max_id + 1)


def _find_existing_supplier_row(ws, email_col: int, email_norm: str):
    for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        raw = str(row[email_col] or "") if row and email_col < len(row) else ""
        if normalize_text(raw) == email_norm:
            return i, row
    return None, None


def upsert_supplier_in_master(config: AppConfig, payload: dict) -> tuple[bool, str, str]:
    """
    Writes supplier and item links to master XLSX with lock.
    Returns (ok, msg, supplier_id)
    """
    xlsx = _xlsx_master_path(config)
    if xlsx is None or not xlsx.exists():
        return False, "mestre_indisponivel", ""

    schema = detect_schema(str(xlsx))
    if schema != "professional":
        return False, "schema_nao_professional", ""

    lock = xlsx.parent / "master_xlsx.lock"
    with _file_lock(lock):
        wb = load_workbook(str(xlsx))
        try:
            if "FORNECEDORES" not in wb.sheetnames or "FORNECEDOR_ITENS" not in wb.sheetnames:
                return False, "abas_obrigatorias_ausentes", ""
            ws_f = wb["FORNECEDORES"]
            ws_l = wb["FORNECEDOR_ITENS"]

            headers_f = [normalize_text(str(c.value or "")) for c in ws_f[3]]
            headers_l = [normalize_text(str(c.value or "")) for c in ws_l[3]]

            def idx_f(name: str, default: int) -> int:
                return headers_f.index(name) if name in headers_f else default

            def idx_l(name: str, default: int) -> int:
                return headers_l.index(name) if name in headers_l else default

            f_supplier_id = idx_f("supplier_id", 0)
            f_empresa = idx_f("empresa", 1)
            f_contato = idx_f("contato", 2)
            f_telefone = idx_f("telefone", 3)
            f_email = idx_f("email", 4)
            f_endereco = idx_f("endereco", 5)
            f_bairro = idx_f("bairro", 6)
            f_cidade = idx_f("cidade", 7)
            f_uf = idx_f("uf", 8)
            f_obs = idx_f("obs", 9)

            l_link_id = idx_l("link_id", 0)
            l_supplier_id = idx_l("supplier_id", 1)
            l_item_id = idx_l("item_id", 2)
            l_serv_corte = idx_l("serv_corte", 3)
            l_serv_dobra = idx_l("serv_dobra", 4)
            l_serv_entrega = idx_l("serv_entrega", 5)
            l_serv_retira = idx_l("serv_retira", 6)

            email_norm = normalize_text(str(payload.get("email") or ""))
            row_idx, row = _find_existing_supplier_row(ws_f, f_email, email_norm)
            if row_idx:
                supplier_id = str(row[f_supplier_id] or "").strip()
                if not supplier_id:
                    supplier_id = _next_numeric_id(ws_f, f_supplier_id)
                    ws_f.cell(row=row_idx, column=f_supplier_id + 1, value=supplier_id)
            else:
                supplier_id = _next_numeric_id(ws_f, f_supplier_id)
                new_row = [""] * max(len(headers_f), 10)
                new_row[f_supplier_id] = supplier_id
                new_row[f_empresa] = payload.get("empresa", "")
                new_row[f_contato] = payload.get("contato", "")
                new_row[f_telefone] = payload.get("telefone", "")
                new_row[f_email] = payload.get("email", "")
                new_row[f_endereco] = payload.get("endereco", "")
                new_row[f_bairro] = ""
                new_row[f_cidade] = payload.get("cidade", "")
                new_row[f_uf] = payload.get("uf", "")
                new_row[f_obs] = payload.get("obs", "")
                ws_f.append(new_row)

            # existing links for supplier
            existing_links = set()
            for r in ws_l.iter_rows(min_row=4, values_only=True):
                if not r:
                    continue
                sid = str(r[l_supplier_id] or "").strip() if l_supplier_id < len(r) else ""
                iid = str(r[l_item_id] or "").strip() if l_item_id < len(r) else ""
                if sid and iid:
                    existing_links.add((sid, iid))

            for item in payload.get("items", []) or []:
                iid = str(item.get("item_id") or "").strip()
                if not iid:
                    continue
                if (supplier_id, iid) in existing_links:
                    continue
                link_id = _next_numeric_id(ws_l, l_link_id)
                row_l = [""] * max(len(headers_l), 7)
                row_l[l_link_id] = link_id
                row_l[l_supplier_id] = supplier_id
                row_l[l_item_id] = iid
                row_l[l_serv_corte] = "SIM" if item.get("serv_corte") else ""
                row_l[l_serv_dobra] = "SIM" if item.get("serv_dobra") else ""
                row_l[l_serv_entrega] = "SIM" if item.get("serv_entrega") else ""
                row_l[l_serv_retira] = "SIM" if item.get("serv_retira") else ""
                ws_l.append(row_l)
                existing_links.add((supplier_id, iid))

            wb.save(str(xlsx))
            try:
                _write_master_sync_marker(xlsx)
            except Exception:
                # Marker is best-effort and must not fail supplier write.
                pass
            return True, "ok", supplier_id
        finally:
            wb.close()
