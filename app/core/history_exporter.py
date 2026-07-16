from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def _clean(value: Any) -> str:
    return str(value or "").strip()


def _autosize_headers(ws) -> None:
    widths: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), len(str(cell.value)))
    for col, width in widths.items():
        ws.column_dimensions[col].width = min(max(width + 2, 10), 80)


def _extra(row: dict, key: str) -> str:
    extra = row.get("extra") or {}
    if not isinstance(extra, dict):
        return ""
    return _clean(extra.get(key))


def _request_text(row: dict) -> str:
    text = _extra(row, "request_text")
    if text:
        return text
    return _clean(row.get("product_query"))


def _account_label(row: dict) -> str:
    sender = _extra(row, "sender").casefold()
    if "@empresa-a." in sender:
        return "Empresa A"
    if "@empresa-b." in sender:
        return "Empresa B"
    if sender:
        return _extra(row, "sender")
    return _clean(row.get("user"))


def _recipient_summary(row: dict) -> str:
    recipients = row.get("recipients") or []
    parts: list[str] = []
    for recipient in recipients:
        if not isinstance(recipient, dict):
            continue
        company = _clean(recipient.get("empresa"))
        contact = _clean(recipient.get("contato_nome"))
        email = _clean(recipient.get("email"))
        if company and contact:
            label = f"{company} ({contact})"
        elif company:
            label = company
        elif contact:
            label = contact
        else:
            label = email
        if email and email not in label:
            label = f"{label} <{email}>"
        if label:
            parts.append(label)
    return " | ".join(parts[:6])


def export_history_rows_to_xlsx(rows: list[dict], output_file: Path) -> Path:
    output_file.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "RESUMO"
    ws_summary.append(
        [
            "DATA_HORA",
            "PEDIDO",
            "SOLICITANTE",
            "LIBERADOR",
            "DESTINATARIOS",
            "CONTA_ENVIO",
            "ASSINATURA",
        ]
    )

    ws_technical = wb.create_sheet("TECNICO")
    ws_technical.append(
        [
            "EVENT_ID",
            "DATA_HORA",
            "TIPO",
            "STATUS",
            "PRODUTO",
            "SOLICITANTE",
            "LIBERADOR",
            "NUM_REQUISICAO",
            "DATA_LIBERACAO",
            "FINALIDADE",
            "SETOR",
            "USUARIO",
            "PC",
            "CONTA_ENVIO",
            "ASSUNTO",
            "CORPO",
            "DESTINATARIOS",
            "ERRO",
        ]
    )

    for row in rows:
        pedido = _request_text(row)
        sender_label = _account_label(row)
        recipients_summary = _recipient_summary(row)
        ws_summary.append(
            [
                _clean(row.get("ts")),
                pedido,
                _extra(row, "requester_name"),
                _extra(row, "approver_name"),
                recipients_summary,
                sender_label,
                _extra(row, "signature_owner"),
            ]
        )

        ws_technical.append(
            [
                _clean(row.get("event_id")),
                _clean(row.get("ts")),
                _clean(row.get("event_type")),
                _clean(row.get("status")),
                _clean(row.get("product_query")),
                _extra(row, "requester_name"),
                _extra(row, "approver_name"),
                _extra(row, "request_number"),
                _extra(row, "release_date"),
                _extra(row, "request_purpose"),
                _extra(row, "request_department"),
                _clean(row.get("user")),
                _clean(row.get("pc_name")),
                sender_label,
                _clean(row.get("subject")),
                _clean(row.get("body")),
                recipients_summary,
                _extra(row, "error"),
            ]
        )

    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    for ws in (ws_summary, ws_technical):
        for c in ws[1]:
            c.font = bold
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        _autosize_headers(ws)

    for ws in (ws_summary, ws_technical):
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = wrap

    wb.save(output_file)
    return output_file
