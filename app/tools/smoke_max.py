from __future__ import annotations

import tempfile
from pathlib import Path

from openpyxl import Workbook

from app.core.catalog_db import CatalogDB
from app.core.config import AppConfig
from app.core.supplier_meta_store_nas import SupplierMeta, SupplierMetaStoreNAS


def _build_mock(path: Path):
    wb = Workbook()
    ws_f = wb.active
    ws_f.title = "FORNECEDORES"
    ws_f.append([""] * 10)
    ws_f.append([""] * 10)
    ws_f.append(["SUPPLIER_ID", "EMPRESA", "CONTATO", "TELEFONE", "EMAIL", "ENDERECO", "BAIRRO", "CIDADE", "UF", "OBS"])
    ws_f.append(["1", "Aço Modelo", "Operador A", "", "contato@aco-modelo.invalid", "", "", "Cidade", "UF", ""])

    ws_i = wb.create_sheet("ITENS")
    ws_i.append([""] * 5)
    ws_i.append([""] * 5)
    ws_i.append(["ITEM_ID", "CATEGORIA", "SUBTIPO", "ITEM", "ATIVO"])
    ws_i.append(["I1", "Chapas", "Inox", "Chapa Inox", "SIM"])

    ws_l = wb.create_sheet("FORNECEDOR_ITENS")
    ws_l.append([""] * 7)
    ws_l.append([""] * 7)
    ws_l.append(["LINK_ID", "SUPPLIER_ID", "ITEM_ID", "SERV_CORTE", "SERV_DOBRA", "SERV_ENTREGA", "SERV_RETIRA"])
    ws_l.append(["1", "1", "I1", "SIM", "", "SIM", ""])

    ws_syn = wb.create_sheet("SINONIMOS")
    ws_syn.append(["termo_base", "alias"])
    ws_syn.append(["chapa inox", "inox chapa"])

    wb.save(path)


def main() -> int:
    with tempfile.TemporaryDirectory() as td:
        xlsx = Path(td) / "mock_max.xlsx"
        _build_mock(xlsx)

        cfg = AppConfig.load()
        cfg.xlsx_sources = [str(xlsx)]
        cfg.nas_master_path = str(xlsx)

        db = CatalogDB(cfg)
        ok_idx, msg_idx = db.rebuild_if_needed(force=True)
        print(f"index_ok={ok_idx} msg={msg_idx}")
        if not ok_idx:
            return 2

        q1 = db.query_suppliers("inox chapa", category="Chapas", limit=50)
        print(f"query_count={len(q1)} top={(q1[0].name if q1 else '-')}")
        if not q1:
            return 2

        cfg_off = AppConfig.load()
        cfg_off.nas_master_path = r"Z:\naoexiste\Cadastro-fornecedores.xlsx"
        m_off = SupplierMetaStoreNAS(cfg_off)
        ok_set, msg_set = m_off.set("email:contato@aco-modelo.invalid", SupplierMeta(supplier_key="email:contato@aco-modelo.invalid", status="ARQUIVADO"))
        print(f"meta_set_online={ok_set} msg={msg_set}")

        m_on = SupplierMetaStoreNAS(cfg)
        ok_sync, msg_sync = m_on.sync_outbox()
        print(f"meta_sync={ok_sync} msg={msg_sync}")
        exists = "email:contato@aco-modelo.invalid" in m_on.list_all()
        print(f"meta_present_after_sync={exists}")

        return 0 if exists else 2


if __name__ == "__main__":
    raise SystemExit(main())
