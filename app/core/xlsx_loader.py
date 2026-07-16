from __future__ import annotations
from pathlib import Path
from typing import List
from openpyxl import load_workbook
from .models import Supplier, SupplierOrigin
from .xlsx_mapping import map_columns
from .utils_text import token_set

def _safe_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()

def load_suppliers_from_xlsx(xlsx_path: str, sheet_name: str = "Fornecedores") -> tuple[list[Supplier], list[str]]:
    """Carrega fornecedores de arquivo XLSX"""
    warnings = []
    p = Path(xlsx_path)
    if not p.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {xlsx_path}")

    # read_only avoids loading full workbook in memory (better for weak PCs).
    wb = load_workbook(xlsx_path, data_only=True, read_only=True, keep_links=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba '{sheet_name}' não encontrada no arquivo {xlsx_path}")
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ValueError(f"Aba '{sheet_name}' está vazia: {xlsx_path}")

    header_list = [ _safe_str(h) for h in header_row ]
    colmap = map_columns(header_list)

    suppliers: List[Supplier] = []
    # Data rows start at row 2 (1-indexed)
    row_idx = 1
    for row in rows_iter:
        row_idx += 1
        empresa = _safe_str(row[colmap.empresa]) if colmap.empresa < len(row) else ""
        material = _safe_str(row[colmap.material_produto]) if colmap.material_produto < len(row) else ""
        email = _safe_str(row[colmap.email]) if colmap.email < len(row) else ""
        telefone = _safe_str(row[colmap.telefone]) if colmap.telefone is not None and colmap.telefone < len(row) else ""
        contato = _safe_str(row[colmap.contato_nome]) if colmap.contato_nome is not None and colmap.contato_nome < len(row) else ""
        endereco = _safe_str(row[colmap.endereco]) if colmap.endereco is not None and colmap.endereco < len(row) else ""
        bairro = _safe_str(row[colmap.bairro_cidade]) if colmap.bairro_cidade is not None and colmap.bairro_cidade < len(row) else ""

        # validate
        is_valid = True
        invalid_reason = ""

        if not empresa:
            is_valid = False
            invalid_reason = f"{p.name}:L{row_idx} - EMPRESA vazia"
            warnings.append(invalid_reason)
        elif not material:
            is_valid = False
            invalid_reason = f"{p.name}:L{row_idx} - MATERIAL/PRODUTO vazio"
            warnings.append(invalid_reason)
        elif "@" not in email:
            is_valid = False
            invalid_reason = f"{p.name}:L{row_idx} - EMAIL inválido '{email}'"
            warnings.append(invalid_reason)

        # Criar supplier mesmo se inválido (para mostrar na UI com visual desabilitado)
        s = Supplier(
            empresa=empresa or "(sem empresa)",
            material_produto=material or "(sem material)",
            email=email or "(sem email)",
            telefone=telefone,
            contato_nome=contato,
            endereco=endereco,
            bairro_cidade=bairro,
            is_valid=is_valid,
            invalid_reason=invalid_reason
        )
        s.origins.append(SupplierOrigin(source_file=str(p), source_row=row_idx))
        # tokens include material + empresa + contato + email
        s.tokens = token_set(material + " " + empresa + " " + contato + " " + email)
        suppliers.append(s)

    wb.close()
    return suppliers, warnings
