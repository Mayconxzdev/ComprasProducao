from __future__ import annotations

import tempfile
from pathlib import Path

from openpyxl import Workbook

from app.core.config import AppConfig
from app.core.data_manager import build_index
from app.core.supplier_meta import SupplierMetaStore, supplier_key_from_obj
from app.core.supplier_scoring import score_supplier
from app.core.synonyms_store import SynonymStore


def _make_mock_xlsx(path: Path):
    wb = Workbook()

    ws_f = wb.active
    ws_f.title = "FORNECEDORES"
    ws_f.append([""] * 9)
    ws_f.append([""] * 9)
    ws_f.append(["SUPPLIER_ID", "EMPRESA", "CONTATO", "TELEFONE", "EMAIL", "ENDERECO", "BAIRRO", "CIDADE", "UF", "OBS"])
    ws_f.append(["S1", "Aço Modelo", "Operador A", "111", "vendas@aco-modelo.invalid", "", "", "Cidade", "UF", ""])
    ws_f.append(["S2", "Metal Modelo", "Operador B", "222", "contato@metal-modelo.invalid", "", "", "Cidade", "UF", ""])

    ws_i = wb.create_sheet("ITENS")
    ws_i.append([""] * 5)
    ws_i.append([""] * 5)
    ws_i.append(["ITEM_ID", "CATEGORIA", "SUBTIPO", "ITEM", "ATIVO"])
    ws_i.append(["I1", "Aco", "Inox", "Chapa Inox", "SIM"])

    ws_l = wb.create_sheet("FORNECEDOR_ITENS")
    ws_l.append([""] * 4)
    ws_l.append([""] * 4)
    ws_l.append(["LINK_ID", "SUPPLIER_ID", "ITEM_ID", "SERV_CORTE"])
    ws_l.append(["L1", "S1", "I1", "SIM"])

    ws_s = wb.create_sheet("SINONIMOS")
    ws_s.append(["termo_base", "alias"])
    ws_s.append(["chapa inox", "inox chapa"])

    wb.save(path)


def main() -> int:
    with tempfile.TemporaryDirectory() as td:
        xlsx = Path(td) / "mock_ultra.xlsx"
        _make_mock_xlsx(xlsx)

        cfg = AppConfig.load()
        cfg.xlsx_sources = [str(xlsx)]
        cfg.nas_master_path = str(xlsx)

        index, res = build_index(cfg.xlsx_sources)
        if res.errors:
            print(f"build_errors={res.errors}")
            return 2

        syn = SynonymStore(cfg)
        variants = syn.expand_query("inox chapa")

        all_suppliers = index.get_all_suppliers() if hasattr(index, "get_all_suppliers") else index.suppliers
        meta = SupplierMetaStore(cfg)

        # archive one supplier and rank
        if len(all_suppliers) > 1:
            meta.set_status(supplier_key_from_obj(all_suppliers[1]), "ARQUIVADO")

        ranked = []
        for s in all_suppliers:
            m = meta.get(supplier_key_from_obj(s))
            sc = score_supplier(s, meta=m, base_match_score=2, exact_item_id="I1")
            if m.status != "ARQUIVADO":
                ranked.append((getattr(s, "empresa", ""), sc.total))
        ranked.sort(key=lambda x: x[1], reverse=True)

        print(f"variants={sorted(list(variants))}")
        print(f"ranked_active={ranked}")
        return 0 if ranked else 2


if __name__ == "__main__":
    raise SystemExit(main())
