from __future__ import annotations

import json
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, Optional

from openpyxl import Workbook, load_workbook

from .config import AppConfig, ensure_app_data_dir
from .utils_text import normalize_text

META_SHEET = "FORNECEDORES_META"
META_HEADER = ["supplier_key", "status", "notes", "tags", "last_used_at", "is_favorite"]


@dataclass
class SupplierMeta:
    supplier_key: str
    status: str = "ATIVO"
    notes: str = ""
    tags: str = ""
    last_used_at: str = ""
    is_favorite: bool = False


class SupplierMetaStore:
    def __init__(self, config: AppConfig):
        self.config = config
        self._cache: Dict[str, SupplierMeta] = {}
        self._loaded = False

    @property
    def fallback_json_path(self) -> Path:
        return ensure_app_data_dir() / "supplier_meta_cache.json"

    def _xlsx_path(self) -> Optional[Path]:
        if self.config.xlsx_sources:
            return Path(self.config.xlsx_sources[0])
        nas = (self.config.nas_master_path or "").strip()
        if nas.lower().endswith((".xlsx", ".xlsm")):
            return Path(nas)
        return None

    def _load_from_json(self) -> Dict[str, SupplierMeta]:
        p = self.fallback_json_path
        if not p.exists():
            return {}
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
            out: Dict[str, SupplierMeta] = {}
            for row in data if isinstance(data, list) else []:
                if not isinstance(row, dict):
                    continue
                key = str(row.get("supplier_key") or "").strip()
                if not key:
                    continue
                out[key] = SupplierMeta(
                    supplier_key=key,
                    status=str(row.get("status") or "ATIVO").upper(),
                    notes=str(row.get("notes") or ""),
                    tags=str(row.get("tags") or ""),
                    last_used_at=str(row.get("last_used_at") or ""),
                    is_favorite=bool(row.get("is_favorite") or False),
                )
            return out
        except Exception:
            return {}

    def _save_json_fallback(self, metas: Dict[str, SupplierMeta]) -> None:
        rows = [asdict(m) for m in metas.values()]
        self.fallback_json_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")

    def _ensure_meta_sheet(self, wb) -> None:
        if META_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(META_SHEET)
            ws.append(META_HEADER)
            return
        ws = wb[META_SHEET]
        if ws.max_row < 1:
            ws.append(META_HEADER)

    def _read_xlsx_meta(self, xlsx_path: Path) -> Dict[str, SupplierMeta]:
        if not xlsx_path.exists():
            return {}
        wb = load_workbook(str(xlsx_path))
        try:
            self._ensure_meta_sheet(wb)
            ws = wb[META_SHEET]
            headers = [str(c.value or "").strip() for c in ws[1]]
            idx = {h: i for i, h in enumerate(headers)}
            out: Dict[str, SupplierMeta] = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                key = str(row[idx.get("supplier_key", 0)] or "").strip() if row else ""
                if not key:
                    continue
                out[key] = SupplierMeta(
                    supplier_key=key,
                    status=str(row[idx.get("status", 1)] or "ATIVO").upper(),
                    notes=str(row[idx.get("notes", 2)] or ""),
                    tags=str(row[idx.get("tags", 3)] or ""),
                    last_used_at=str(row[idx.get("last_used_at", 4)] or ""),
                    is_favorite=str(row[idx.get("is_favorite", 5)] or "").strip().upper() in {"1", "TRUE", "SIM", "X"},
                )
            return out
        finally:
            wb.close()

    def _write_xlsx_meta(self, xlsx_path: Path, metas: Dict[str, SupplierMeta]) -> None:
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        if xlsx_path.exists():
            wb = load_workbook(str(xlsx_path))
        else:
            wb = Workbook()
            wb.remove(wb.active)
        try:
            self._ensure_meta_sheet(wb)
            ws = wb[META_SHEET]
            ws.delete_rows(2, ws.max_row)
            for meta in metas.values():
                ws.append([
                    meta.supplier_key,
                    meta.status,
                    meta.notes,
                    meta.tags,
                    meta.last_used_at,
                    "1" if meta.is_favorite else "0",
                ])
            wb.save(str(xlsx_path))
        finally:
            wb.close()

    def load(self, force: bool = False) -> Dict[str, SupplierMeta]:
        if self._loaded and not force:
            return self._cache
        xlsx = self._xlsx_path()
        metas: Dict[str, SupplierMeta] = {}
        if xlsx is not None:
            try:
                metas = self._read_xlsx_meta(xlsx)
            except Exception:
                metas = {}
        if not metas:
            metas = self._load_from_json()
        self._cache = metas
        self._loaded = True
        return self._cache

    def save(self) -> None:
        metas = self._cache
        xlsx = self._xlsx_path()
        if xlsx is not None:
            try:
                self._write_xlsx_meta(xlsx, metas)
                self._save_json_fallback(metas)
                return
            except Exception:
                pass
        self._save_json_fallback(metas)

    def get(self, supplier_key: str) -> SupplierMeta:
        self.load()
        key = (supplier_key or "").strip()
        if key not in self._cache:
            self._cache[key] = SupplierMeta(supplier_key=key)
        return self._cache[key]

    def set_status(self, supplier_key: str, status: str) -> None:
        meta = self.get(supplier_key)
        meta.status = (status or "ATIVO").upper()
        self.save()

    def set_favorite(self, supplier_key: str, value: bool) -> None:
        meta = self.get(supplier_key)
        meta.is_favorite = bool(value)
        self.save()

    def touch_last_used(self, supplier_key: str) -> None:
        meta = self.get(supplier_key)
        meta.last_used_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.save()

    def bulk_set_status(self, supplier_keys: Iterable[str], status: str) -> int:
        count = 0
        for key in supplier_keys:
            if not key:
                continue
            self.get(key).status = status.upper()
            count += 1
        if count:
            self.save()
        return count


def supplier_key_from_obj(supplier) -> str:
    sid = getattr(supplier, "supplier_id", "") or ""
    if sid:
        return f"id:{normalize_text(sid)}"
    email = normalize_text(getattr(supplier, "email", "") or "")
    if email:
        return f"email:{email}"
    empresa = normalize_text(getattr(supplier, "empresa", "") or "")
    cidade = normalize_text(getattr(supplier, "cidade", "") or getattr(supplier, "bairro_cidade", "") or "")
    uf = normalize_text(getattr(supplier, "uf", "") or "")
    return f"name:{empresa}|{cidade}|{uf}"
