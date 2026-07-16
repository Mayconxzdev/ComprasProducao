"""
Professional XLSX loader with multi-sheet support and manual JOINs

Handles complex Excel files with:
- FORNECEDORES sheet (suppliers)
- ITENS sheet (items/products)
- FORNECEDOR_ITENS sheet (supplier-item links)
- TERMOS sheet (synonyms)
"""
from __future__ import annotations
import logging
from pathlib import Path
from typing import Dict, List
from openpyxl import load_workbook
from .models_pro import Item, SupplierItemLink, SupplierPro, Synonym, ProDataSet
from .utils_text import normalize_text, token_set

logger = logging.getLogger(__name__)

def _safe_str(v) -> str:
    """Safely convert cell value to string"""
    if v is None:
        return ""
    return str(v).strip()

def _safe_bool(v) -> bool:
    """Convert cell value to boolean"""
    if v is None:
        return False
    s = str(v).strip().upper()
    return s in ('SIM', 'TRUE', '1', 'X', 'YES')

def _safe_int(v) -> int:
    """Safely convert to int"""
    if v is None:
        return 0
    try:
        return int(float(v))
    except:
        return 0

def detect_schema(xlsx_path: str) -> str:
    """
    Detect if XLSX is 'professional' (multi-sheet) or 'simple' (single sheet)

    Returns: 'professional' or 'simple'
    """
    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        # Professional schema requires these sheets
        required_sheets = ['FORNECEDORES', 'ITENS', 'FORNECEDOR_ITENS']
        if all(sheet in sheet_names for sheet in required_sheets):
            logger.info(f"Detected PROFESSIONAL schema in {xlsx_path}")
            return 'professional'

        logger.info(f"Detected SIMPLE schema in {xlsx_path}")
        return 'simple'
    except Exception as e:
        logger.warning(f"Error detecting schema: {e}, defaulting to simple")
        return 'simple'

def load_professional_xlsx(xlsx_path: str) -> ProDataSet:
    """
    Load professional multi-sheet XLSX and perform manual JOINs

    Expected sheets:
    - FORNECEDORES: supplier data (header row 3, data row 4+)
    - ITENS: item/product catalog (header row 3, data row 4+)
    - FORNECEDOR_ITENS: supplier-item links (header row 3, data row 4+)
    - TERMOS: synonyms (optional)

    Returns: ProDataSet with suppliers, items, links, synonyms
    """
    p = Path(xlsx_path)
    if not p.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {xlsx_path}")

    logger.info(f"Loading professional XLSX: {xlsx_path}")
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)

    dataset = ProDataSet(
        suppliers={},
        items={},
        links=[],
        synonyms=[],
        warnings=[]
    )

    try:
        # Load ITENS first (needed for validation)
        if 'ITENS' in wb.sheetnames:
            _load_items_sheet(wb['ITENS'], dataset, str(p))
        else:
            dataset.warnings.append("Aba ITENS não encontrada")

        # Load FORNECEDORES
        if 'FORNECEDORES' in wb.sheetnames:
            _load_fornecedores_sheet(wb['FORNECEDORES'], dataset, str(p))
        else:
            raise ValueError("Aba FORNECEDORES não encontrada")

        # Load FORNECEDOR_ITENS (links)
        if 'FORNECEDOR_ITENS' in wb.sheetnames:
            _load_fornecedor_itens_sheet(wb['FORNECEDOR_ITENS'], dataset, str(p))
        else:
            dataset.warnings.append("Aba FORNECEDOR_ITENS não encontrada")

        # Load TERMOS (synonyms)
        if 'TERMOS' in wb.sheetnames:
            _load_termos_sheet(wb['TERMOS'], dataset)

        # Perform JOINs: link suppliers with items
        _perform_joins(dataset)

        # Compute search tokens for each supplier
        _compute_tokens(dataset)

    finally:
        wb.close()

    logger.info(f"Loaded {len(dataset.suppliers)} suppliers, {len(dataset.items)} items, "
                f"{len(dataset.links)} links, {len(dataset.synonyms)} synonyms")

    return dataset

def _load_items_sheet(ws, dataset: ProDataSet, source_file: str):
    """Load ITENS sheet (header row 3, data row 4+)"""
    logger.info("Loading ITENS sheet...")

    # Skip to header row (row 3)
    rows_iter = ws.iter_rows(min_row=3, values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        dataset.warnings.append("ITENS sheet vazia")
        return

    # Map columns (expected: ITEM_ID, CATEGORIA, SUBTIPO, ITEM, ATIVO)
    col_map = _map_item_columns([_safe_str(h) for h in header_row])

    row_idx = 3
    for row in rows_iter:
        row_idx += 1

        item_id = _safe_str(row[col_map.get('item_id', 0)] if col_map.get('item_id', 0) < len(row) else None)
        if not item_id:
            continue  # Skip rows without item_id

        categoria = _safe_str(row[col_map.get('categoria', 1)] if col_map.get('categoria', 1) < len(row) else None)
        subtipo = _safe_str(row[col_map.get('subtipo', 2)] if col_map.get('subtipo', 2) < len(row) else None)
        item = _safe_str(row[col_map.get('item', 3)] if col_map.get('item', 3) < len(row) else None)
        ativo = _safe_bool(row[col_map.get('ativo', 4)] if col_map.get('ativo', 4) < len(row) else None)

        dataset.items[item_id] = Item(
            item_id=item_id,
            categoria=categoria,
            subtipo=subtipo,
            item=item,
            ativo=ativo
        )

    logger.info(f"Loaded {len(dataset.items)} items")

def _load_fornecedores_sheet(ws, dataset: ProDataSet, source_file: str):
    """Load FORNECEDORES sheet (header row 3, data row 4+)"""
    logger.info("Loading FORNECEDORES sheet...")

    rows_iter = ws.iter_rows(min_row=3, values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ValueError("FORNECEDORES sheet vazia")

    col_map = _map_fornecedores_columns([_safe_str(h) for h in header_row])

    row_idx = 3
    for row in rows_iter:
        row_idx += 1

        supplier_id = _safe_str(row[col_map.get('supplier_id', 0)] if col_map.get('supplier_id', 0) < len(row) else None)
        if not supplier_id:
            continue

        empresa = _safe_str(row[col_map.get('empresa', 1)] if col_map.get('empresa', 1) < len(row) else None)
        contato = _safe_str(row[col_map.get('contato', 2)] if col_map.get('contato', 2) < len(row) else None)
        telefone = _safe_str(row[col_map.get('telefone', 3)] if col_map.get('telefone', 3) < len(row) else None)
        email = _safe_str(row[col_map.get('email', 4)] if col_map.get('email', 4) < len(row) else None)
        endereco = _safe_str(row[col_map.get('endereco', 5)] if col_map.get('endereco', 5) < len(row) else None)
        bairro = _safe_str(row[col_map.get('bairro', 6)] if col_map.get('bairro', 6) < len(row) else None)
        cidade = _safe_str(row[col_map.get('cidade', 7)] if col_map.get('cidade', 7) < len(row) else None)
        uf = _safe_str(row[col_map.get('uf', 8)] if col_map.get('uf', 8) < len(row) else None)
        obs = _safe_str(row[col_map.get('obs', 9)] if col_map.get('obs', 9) < len(row) else None)

        # Validation
        is_valid = True
        invalid_reason = ""

        if not empresa:
            is_valid = False
            invalid_reason = f"{Path(source_file).name}:L{row_idx} - EMPRESA vazia"
        elif "@" not in email:
            is_valid = False
            invalid_reason = f"{Path(source_file).name}:L{row_idx} - EMAIL inválido"

        supplier = SupplierPro(
            supplier_id=supplier_id,
            empresa=empresa or "(sem empresa)",
            contato=contato,
            telefone=telefone,
            email=email,
            endereco=endereco,
            bairro=bairro,
            cidade=cidade,
            uf=uf,
            obs=obs,
            is_valid=is_valid,
            invalid_reason=invalid_reason,
            source_file=source_file
        )

        dataset.suppliers[supplier_id] = supplier

        if not is_valid:
            dataset.warnings.append(invalid_reason)

    logger.info(f"Loaded {len(dataset.suppliers)} suppliers")

def _load_fornecedor_itens_sheet(ws, dataset: ProDataSet, source_file: str):
    """Load FORNECEDOR_ITENS sheet (header row 3, data row 4+)"""
    logger.info("Loading FORNECEDOR_ITENS sheet...")

    rows_iter = ws.iter_rows(min_row=3, values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        dataset.warnings.append("FORNECEDOR_ITENS sheet vazia")
        return

    col_map = _map_fornecedor_itens_columns([_safe_str(h) for h in header_row])

    row_idx = 3
    for row in rows_iter:
        row_idx += 1

        link_id = _safe_str(row[col_map.get('link_id', 0)] if col_map.get('link_id', 0) < len(row) else None)
        supplier_id = _safe_str(row[col_map.get('supplier_id', 1)] if col_map.get('supplier_id', 1) < len(row) else None)
        item_id = _safe_str(row[col_map.get('item_id', 2)] if col_map.get('item_id', 2) < len(row) else None)

        if not supplier_id or not item_id:
            continue

        serv_corte = _safe_bool(row[col_map.get('serv_corte', 3)] if col_map.get('serv_corte', 3) < len(row) else None)
        serv_dobra = _safe_bool(row[col_map.get('serv_dobra', 4)] if col_map.get('serv_dobra', 4) < len(row) else None)
        serv_entrega = _safe_bool(row[col_map.get('serv_entrega', 5)] if col_map.get('serv_entrega', 5) < len(row) else None)
        serv_retira = _safe_bool(row[col_map.get('serv_retira', 6)] if col_map.get('serv_retira', 6) < len(row) else None)
        obs_link = _safe_str(row[col_map.get('obs_link', 7)] if col_map.get('obs_link', 7) < len(row) else None)
        prioridade = _safe_int(row[col_map.get('prioridade', 8)] if col_map.get('prioridade', 8) < len(row) else None)
        prazo = _safe_int(row[col_map.get('prazo_entrega_dias', 9)] if col_map.get('prazo_entrega_dias', 9) < len(row) else None)

        link = SupplierItemLink(
            link_id=link_id or f"{supplier_id}_{item_id}",
            supplier_id=supplier_id,
            item_id=item_id,
            serv_corte=serv_corte,
            serv_dobra=serv_dobra,
            serv_entrega=serv_entrega,
            serv_retira=serv_retira,
            obs_link=obs_link,
            prioridade=prioridade,
            prazo_entrega_dias=prazo
        )

        dataset.links.append(link)

    logger.info(f"Loaded {len(dataset.links)} supplier-item links")

def _load_termos_sheet(ws, dataset: ProDataSet):
    """Load TERMOS sheet for synonyms"""
    logger.info("Loading TERMOS sheet...")

    # TERMOS may have different structure, check first row
    rows_iter = ws.iter_rows(values_only=True)

    for row in rows_iter:
        if not row or len(row) < 2:
            continue

        tipo = _safe_str(row[0]) if len(row) > 0 else ""
        termo = _safe_str(row[1]) if len(row) > 1 else ""
        item_id = _safe_str(row[2]) if len(row) > 2 else ""

        # Only load synonyms
        if tipo.upper() == "SINONIMO" and termo and item_id:
            dataset.synonyms.append(Synonym(termo=termo, item_id=item_id))

    logger.info(f"Loaded {len(dataset.synonyms)} synonyms")

def _perform_joins(dataset: ProDataSet):
    """Perform manual JOINs between suppliers and items"""
    logger.info("Performing JOINs...")

    for link in dataset.links:
        # Get supplier
        supplier = dataset.suppliers.get(link.supplier_id)
        if not supplier:
            continue

        # Get item
        item = dataset.items.get(link.item_id)
        if not item:
            continue

        # Add item to supplier
        if item not in supplier.items:
            supplier.items.append(item)

        # Add link to supplier
        supplier.links.append(link)

        # Aggregate services (OR logic - if ANY link has service, supplier has it)
        if link.serv_corte:
            supplier.serv_corte = True
        if link.serv_dobra:
            supplier.serv_dobra = True
        if link.serv_entrega:
            supplier.serv_entrega = True
        if link.serv_retira:
            supplier.serv_retira = True

def _compute_tokens(dataset: ProDataSet):
    """Compute search tokens for each supplier"""
    for supplier in dataset.suppliers.values():
        # Combine all text fields
        text_parts = [
            supplier.empresa,
            supplier.contato,
            supplier.cidade,
            supplier.uf,
            supplier.email
        ]

        # Add item names
        for item in supplier.items:
            text_parts.extend([item.categoria, item.subtipo, item.item])

        # Compute tokens
        combined_text = " ".join(filter(None, text_parts))
        supplier.tokens = token_set(combined_text)

def _map_item_columns(headers: List[str]) -> Dict[str, int]:
    """Map ITENS columns by fuzzy matching"""
    col_map = {}
    for idx, h in enumerate(headers):
        h_norm = normalize_text(h)
        if 'item_id' in h_norm or 'itemid' in h_norm:
            col_map['item_id'] = idx
        elif 'categoria' in h_norm:
            col_map['categoria'] = idx
        elif 'subtipo' in h_norm:
            col_map['subtipo'] = idx
        elif h_norm == 'item':
            col_map['item'] = idx
        elif 'ativo' in h_norm or 'ativa' in h_norm:
            col_map['ativo'] = idx
    return col_map

def _map_fornecedores_columns(headers: List[str]) -> Dict[str, int]:
    """Map FORNECEDORES columns"""
    col_map = {}
    for idx, h in enumerate(headers):
        h_norm = normalize_text(h)
        if 'supplier_id' in h_norm or 'supplierid' in h_norm or 'fornecedor_id' in h_norm:
            col_map['supplier_id'] = idx
        elif 'empresa' in h_norm:
            col_map['empresa'] = idx
        elif 'contato' in h_norm:
            col_map['contato'] = idx
        elif 'telefone' in h_norm or 'fone' in h_norm:
            col_map['telefone'] = idx
        elif 'email' in h_norm or 'mail' in h_norm:
            col_map['email'] = idx
        elif 'endereco' in h_norm:
            col_map['endereco'] = idx
        elif 'bairro' in h_norm:
            col_map['bairro'] = idx
        elif 'cidade' in h_norm:
            col_map['cidade'] = idx
        elif h_norm == 'uf' or 'estado' in h_norm:
            col_map['uf'] = idx
        elif 'obs' in h_norm or 'observa' in h_norm:
            col_map['obs'] = idx
    return col_map

def _map_fornecedor_itens_columns(headers: List[str]) -> Dict[str, int]:
    """Map FORNECEDOR_ITENS columns (use ONLY non-formula columns)"""
    col_map = {}
    for idx, h in enumerate(headers):
        h_norm = normalize_text(h)
        if 'link_id' in h_norm or 'linkid' in h_norm:
            col_map['link_id'] = idx
        elif 'supplier_id' in h_norm or 'fornecedor_id' in h_norm:
            col_map['supplier_id'] = idx
        elif 'item_id' in h_norm:
            col_map['item_id'] = idx
        elif 'serv_corte' in h_norm or 'servico_corte' in h_norm:
            col_map['serv_corte'] = idx
        elif 'serv_dobra' in h_norm or 'servico_dobra' in h_norm:
            col_map['serv_dobra'] = idx
        elif 'serv_entrega' in h_norm or 'servico_entrega' in h_norm:
            col_map['serv_entrega'] = idx
        elif 'serv_retira' in h_norm or 'servico_retira' in h_norm:
            col_map['serv_retira'] = idx
        elif 'obs_link' in h_norm or 'obslink' in h_norm:
            col_map['obs_link'] = idx
        elif 'prioridade' in h_norm:
            col_map['prioridade'] = idx
        elif 'prazo' in h_norm:
            col_map['prazo_entrega_dias'] = idx
    return col_map
