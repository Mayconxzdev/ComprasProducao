from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import List

from openpyxl import load_workbook

from .config import AppConfig


@dataclass
class ValidationIssue:
    index: int
    field: str
    message: str


@dataclass
class ValidationResult:
    ok: bool
    issues: List[ValidationIssue]


@dataclass
class CategoryRule:
    categoria: str
    required_fields: List[str]
    help_text: str


def _extract_qty(line: str) -> bool:
    text = str(line or "").strip()
    if not text:
        return False
    if _QTY_PREFIX_RE.match(text):
        return True

    # Legacy fallback: qty before hyphen.
    left = text.split("-", 1)[0].strip()
    return bool(left and any(ch.isdigit() for ch in left) and "-" in text)


_QTY_PREFIX_RE = re.compile(r"^\s*(\d+)(?:\s*[A-Za-zÀ-ÿ]{1,6})?\s*(?:-|\|)\s*(.+)$")


def _line_body(line: str) -> str:
    text = str(line or "").strip()
    if not text:
        return ""

    m = _QTY_PREFIX_RE.match(text)
    if m:
        return str(m.group(2) or "").strip()
    if "-" in text:
        return text.split("-", 1)[1].strip()
    return text


def _pipe_parts(line: str) -> List[str]:
    body = _line_body(line)
    return [p.strip() for p in body.split("|") if p.strip()]


def _extract_unit(line: str) -> bool:
    body = _line_body(line)
    parts = [p.strip() for p in body.split("-") if p.strip()]
    if len(parts) >= 3:
        return True
    pipe_parts = _pipe_parts(line)
    if len(pipe_parts) >= 4:
        return True
    tokens = (line or "").lower().split()
    for t in tokens:
        if t in {"kg", "un", "pc", "pcs", "m", "mm", "cm", "mt", "barra", "chapas"}:
            return True
    return False


def _extract_description(line: str) -> bool:
    body = _line_body(line)
    parts = [p.strip() for p in body.split("-") if p.strip()]
    if len(parts) < 2:
        pipe_parts = _pipe_parts(line)
        if not pipe_parts:
            return False
        return len(pipe_parts[0]) >= 2
    return len(parts[1]) >= 2


def load_category_rules(config: AppConfig) -> List[CategoryRule]:
    xlsx = None
    if config.xlsx_sources:
        xlsx = Path(config.xlsx_sources[0])
    elif str(config.nas_master_path or "").lower().endswith((".xlsx", ".xlsm")):
        xlsx = Path(config.nas_master_path)
    if xlsx is None or not xlsx.exists():
        return []

    wb = load_workbook(str(xlsx), data_only=True, read_only=True)
    try:
        if "REGRAS_CATEGORIA" not in wb.sheetnames:
            return []
        ws = wb["REGRAS_CATEGORIA"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(h or "").strip().lower() for h in rows[0]]
        def _idx(name: str, default: int) -> int:
            try:
                return header.index(name)
            except ValueError:
                return default
        c_idx = _idx("categoria", 0)
        f_idx = _idx("campos_obrigatorios", 1)
        h_idx = _idx("texto_ajuda", 2)

        out: List[CategoryRule] = []
        for row in rows[1:]:
            if not row:
                continue
            cat = str(row[c_idx] or "").strip().lower() if c_idx < len(row) else ""
            fields = str(row[f_idx] or "").strip().lower() if f_idx < len(row) else ""
            help_text = str(row[h_idx] or "").strip() if h_idx < len(row) else ""
            if not cat:
                continue
            out.append(CategoryRule(categoria=cat, required_fields=[f.strip() for f in fields.split(",") if f.strip()], help_text=help_text))
        return out
    finally:
        wb.close()


def validate_quote_items(lines: List[str], *, category_hint: str = "", rules: List[CategoryRule] | None = None) -> ValidationResult:
    issues: List[ValidationIssue] = []
    for i, line in enumerate(lines, start=1):
        if not _extract_qty(line):
            issues.append(ValidationIssue(index=i, field="quantidade", message=f"Item {i}: quantidade ausente"))
        if not _extract_description(line):
            issues.append(ValidationIssue(index=i, field="descricao", message=f"Item {i}: descricao ausente"))
        if not _extract_unit(line):
            issues.append(ValidationIssue(index=i, field="unidade", message=f"Item {i}: unidade/medida ausente"))

    category_hint_norm = (category_hint or "").strip().lower()
    for rule in (rules or []):
        if rule.categoria and rule.categoria in category_hint_norm:
            for i, line in enumerate(lines, start=1):
                lower = (line or "").lower()
                for field in rule.required_fields:
                    f = field.strip().lower()
                    if not f:
                        continue
                    if f == "quantidade" and not _extract_qty(line):
                        continue
                    if f == "descricao" and not _extract_description(line):
                        continue
                    if f in {"unidade", "medida"} and not _extract_unit(line):
                        continue
                    if f not in {"quantidade", "descricao", "unidade", "medida"} and f not in lower:
                        msg = f"Item {i}: obrigatorio '{f}' para categoria '{rule.categoria}'"
                        if rule.help_text:
                            msg += f" ({rule.help_text})"
                        issues.append(ValidationIssue(index=i, field=f, message=msg))

    return ValidationResult(ok=not issues, issues=issues)
