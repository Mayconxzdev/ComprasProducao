from __future__ import annotations
import datetime as _dt
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font
from .history_db import connect, init_db, list_quotes, get_quote


def _autosize_headers(ws) -> None:
    widths = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), len(str(cell.value)))
    for col, width in widths.items():
        ws.column_dimensions[col].width = min(max(width + 2, 10), 60)


def export_history_xlsx(output_dir: str) -> Path:
    """
    Exporta histórico SQLite para XLSX com 3 abas:
    - COTACOES
    - ITENS
    - DESTINATARIOS
    """
    out_dir = Path(output_dir).expanduser()
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = _dt.datetime.now().strftime("%Y-%m-%d_%H%M")
    out_path = out_dir / f"historico_cotacoes_{ts}.xlsx"

    conn = connect()
    try:
        init_db(conn)
        quote_rows = list_quotes(conn, limit=100000)

        wb = Workbook()

        ws_quotes = wb.active
        ws_quotes.title = "COTACOES"
        ws_quotes.append([
            "ID", "DATA_HORA", "PRODUTO_BUSCA", "STATUS", "DESTINATARIOS",
            "USUARIO_PC", "ASSUNTO", "CORPO"
        ])

        ws_items = wb.create_sheet("ITENS")
        ws_items.append(["QUOTE_ID", "ORDEM", "ITEM"])

        ws_recipients = wb.create_sheet("DESTINATARIOS")
        ws_recipients.append([
            "QUOTE_ID", "EMPRESA", "CONTATO", "EMAIL", "TELEFONE",
            "MATERIAL_PRODUTO", "SOURCE_FILE", "SOURCE_ROW"
        ])

        for (qid, _created_at, _product_query, _rc, _status) in quote_rows:
            q = get_quote(conn, qid)
            ws_quotes.append([
                q.id, q.created_at, q.product_query, q.status, len(q.recipients),
                q.user_pc, q.subject, q.body
            ])

            for idx, item in enumerate(q.items, start=1):
                ws_items.append([q.id, idx, item.line_text])

            for r in q.recipients:
                ws_recipients.append([
                    q.id, r.empresa, r.contato_nome, r.email, r.telefone,
                    r.material_produto, r.source_file, r.source_row
                ])

        bold = Font(bold=True)
        for ws in (ws_quotes, ws_items, ws_recipients):
            for c in ws[1]:
                c.font = bold
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            _autosize_headers(ws)

        wb.save(out_path)
        return out_path
    finally:
        conn.close()
