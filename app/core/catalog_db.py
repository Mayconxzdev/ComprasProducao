from __future__ import annotations

import hashlib
import re
import sqlite3
import threading
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Optional

from .config import AppConfig, ensure_app_data_dir
from .perf_metrics import record_timing
from .utils_text import normalize_text
from .xlsx_loader_pro import detect_schema, load_professional_xlsx
from .xlsx_loader import load_suppliers_from_xlsx
from .supplier_meta_store_nas import supplier_key_from_obj

DB_NAME = "catalog.sqlite3"
SCHEMA_VERSION = 2
WORD_RE = re.compile(r"[a-z0-9]+")


@dataclass
class CatalogSupplierRow:
    supplier_key: str
    supplier_id: str
    name: str
    email: str
    city: str
    uf: str
    category: str
    base_score: int


class CatalogDB:
    def __init__(self, config: AppConfig):
        self.config = config
        self.db_path = ensure_app_data_dir() / DB_NAME
        self._lock = threading.Lock()
        self._last_progress = ""

    def connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(str(self.db_path), timeout=30)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA synchronous=NORMAL;")
        conn.execute("PRAGMA temp_store=MEMORY;")
        conn.execute("PRAGMA cache_size=-10000;")
        return conn

    def init_schema(self) -> None:
        with self.connect() as conn:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS suppliers (
                    supplier_key TEXT PRIMARY KEY,
                    supplier_id TEXT,
                    name TEXT,
                    name_norm TEXT,
                    email TEXT,
                    email_norm TEXT,
                    city TEXT,
                    uf TEXT,
                    raw_category TEXT
                );

                CREATE TABLE IF NOT EXISTS items (
                    item_id TEXT PRIMARY KEY,
                    name TEXT,
                    name_norm TEXT,
                    category TEXT
                );

                CREATE TABLE IF NOT EXISTS supplier_items (
                    supplier_key TEXT,
                    item_id TEXT,
                    PRIMARY KEY (supplier_key, item_id)
                );

                CREATE TABLE IF NOT EXISTS synonyms (
                    term_base TEXT,
                    alias_norm TEXT,
                    PRIMARY KEY (term_base, alias_norm)
                );

                CREATE TABLE IF NOT EXISTS meta (
                    k TEXT PRIMARY KEY,
                    v TEXT
                );

                CREATE INDEX IF NOT EXISTS idx_suppliers_name_norm ON suppliers(name_norm);
                CREATE INDEX IF NOT EXISTS idx_suppliers_email_norm ON suppliers(email_norm);
                CREATE INDEX IF NOT EXISTS idx_items_name_norm ON items(name_norm);
                CREATE INDEX IF NOT EXISTS idx_items_category ON items(category);
                CREATE INDEX IF NOT EXISTS idx_supplier_items_item ON supplier_items(item_id);
                CREATE INDEX IF NOT EXISTS idx_supplier_items_supplier ON supplier_items(supplier_key);
                CREATE INDEX IF NOT EXISTS idx_syn_alias ON synonyms(alias_norm);

                CREATE VIRTUAL TABLE IF NOT EXISTS supplier_fts USING fts5(
                    supplier_key UNINDEXED,
                    name,
                    email,
                    category,
                    items,
                    tokenize='unicode61 remove_diacritics 2'
                );
                """
            )
            self._set_meta(conn, "schema_version", str(SCHEMA_VERSION))

    def _set_meta(self, conn: sqlite3.Connection, k: str, v: str) -> None:
        conn.execute("INSERT INTO meta(k,v) VALUES(?,?) ON CONFLICT(k) DO UPDATE SET v=excluded.v", (k, v))

    def _get_meta(self, conn: sqlite3.Connection, k: str) -> str:
        row = conn.execute("SELECT v FROM meta WHERE k=?", (k,)).fetchone()
        return str(row[0]) if row else ""

    def xlsx_path(self) -> Optional[Path]:
        if self.config.xlsx_sources:
            return Path(self.config.xlsx_sources[0])
        p = (self.config.nas_master_path or "").strip()
        if p.lower().endswith((".xlsx", ".xlsm")):
            return Path(p)
        return None

    def fingerprint(self, path: Path) -> str:
        st = path.stat()
        raw = f"{path}|{st.st_size}|{st.st_mtime_ns}".encode("utf-8", errors="ignore")
        return hashlib.sha1(raw).hexdigest()

    def needs_reindex(self, xlsx_path: Path) -> bool:
        self.init_schema()
        fp = self.fingerprint(xlsx_path)
        with self.connect() as conn:
            current = self._get_meta(conn, "xlsx_fingerprint")
            return fp != current

    def rebuild_if_needed(self, *, on_progress: Optional[Callable[[str], None]] = None, force: bool = False) -> tuple[bool, str]:
        self.init_schema()
        xlsx = self.xlsx_path()
        if xlsx is None:
            return False, "sem_xlsx"
        if not xlsx.exists():
            return False, "xlsx_inexistente"
        if not force and not self.needs_reindex(xlsx):
            return True, "catalogo_em_dia"
        return self.rebuild_from_xlsx(xlsx, on_progress=on_progress)

    def rebuild_from_xlsx(self, xlsx_path: Path, *, on_progress: Optional[Callable[[str], None]] = None) -> tuple[bool, str]:
        def p(msg: str):
            self._last_progress = msg
            if on_progress:
                on_progress(msg)

        with self._lock:
            self.init_schema()
            schema = detect_schema(str(xlsx_path))
            p(f"Indexando catalogo ({schema})...")
            try:
                with self.connect() as conn:
                    conn.execute("BEGIN")
                    conn.execute("DELETE FROM supplier_items")
                    conn.execute("DELETE FROM suppliers")
                    conn.execute("DELETE FROM items")
                    conn.execute("DELETE FROM synonyms")

                    if schema == "professional":
                        ds = load_professional_xlsx(str(xlsx_path))
                        for s in ds.suppliers.values():
                            skey = supplier_key_from_obj(s)
                            raw_cat = ", ".join(sorted({str(getattr(i, "categoria", "") or "") for i in (getattr(s, "items", []) or []) if getattr(i, "categoria", "")}))
                            conn.execute(
                                """
                                INSERT OR REPLACE INTO suppliers(
                                  supplier_key, supplier_id, name, name_norm, email, email_norm, city, uf, raw_category
                                ) VALUES(?,?,?,?,?,?,?,?,?)
                                """,
                                (
                                    skey,
                                    str(getattr(s, "supplier_id", "") or ""),
                                    str(getattr(s, "empresa", "") or ""),
                                    normalize_text(str(getattr(s, "empresa", "") or "")),
                                    str(getattr(s, "email", "") or ""),
                                    normalize_text(str(getattr(s, "email", "") or "")),
                                    str(getattr(s, "cidade", "") or ""),
                                    str(getattr(s, "uf", "") or ""),
                                    raw_cat,
                                ),
                            )
                            for it in (getattr(s, "items", []) or []):
                                item_id = str(getattr(it, "item_id", "") or "")
                                if not item_id:
                                    continue
                                conn.execute(
                                    "INSERT OR IGNORE INTO supplier_items(supplier_key,item_id) VALUES(?,?)",
                                    (skey, item_id),
                                )

                        for it in ds.items.values():
                            conn.execute(
                                "INSERT OR REPLACE INTO items(item_id,name,name_norm,category) VALUES(?,?,?,?)",
                                (
                                    str(it.item_id),
                                    str(it.item),
                                    normalize_text(str(it.item)),
                                    str(it.categoria or ""),
                                ),
                            )

                        # TERMOS from professional loader
                        for syn in ds.synonyms:
                            base = normalize_text(str(getattr(syn, "termo", "") or ""))
                            if base:
                                conn.execute(
                                    "INSERT OR IGNORE INTO synonyms(term_base,alias_norm) VALUES(?,?)",
                                    (base, base),
                                )

                    else:
                        suppliers, _warnings = load_suppliers_from_xlsx(str(xlsx_path), sheet_name=self.config.xlsx_sheet_name)
                        for s in suppliers:
                            skey = supplier_key_from_obj(s)
                            conn.execute(
                                """
                                INSERT OR REPLACE INTO suppliers(
                                  supplier_key, supplier_id, name, name_norm, email, email_norm, city, uf, raw_category
                                ) VALUES(?,?,?,?,?,?,?,?,?)
                                """,
                                (
                                    skey,
                                    "",
                                    str(getattr(s, "empresa", "") or ""),
                                    normalize_text(str(getattr(s, "empresa", "") or "")),
                                    str(getattr(s, "email", "") or ""),
                                    normalize_text(str(getattr(s, "email", "") or "")),
                                    str(getattr(s, "bairro_cidade", "") or ""),
                                    "",
                                    str(getattr(s, "material_produto", "") or ""),
                                ),
                            )

                    # Load optional SINONIMOS sheet from xlsx.
                    try:
                        from openpyxl import load_workbook

                        wb = load_workbook(str(xlsx_path), data_only=True, read_only=True)
                        try:
                            if "SINONIMOS" in wb.sheetnames:
                                ws = wb["SINONIMOS"]
                                rows = list(ws.iter_rows(values_only=True))
                                if rows:
                                    h = [normalize_text(str(x or "")) for x in rows[0]]
                                    bi = h.index("termo_base") if "termo_base" in h else 0
                                    ai = h.index("alias") if "alias" in h else 1
                                    for row in rows[1:]:
                                        if not row:
                                            continue
                                        base = normalize_text(str(row[bi] or "")) if bi < len(row) else ""
                                        alias = normalize_text(str(row[ai] or "")) if ai < len(row) else ""
                                        if base and alias:
                                            conn.execute(
                                                "INSERT OR IGNORE INTO synonyms(term_base,alias_norm) VALUES(?,?)",
                                                (base, alias),
                                            )
                        finally:
                            wb.close()
                    except Exception:
                        pass

                    # Índice FTS5 materializado: usado como pré-filtro rápido para
                    # abrir Fornecedores e buscar sem varrer a planilha em tempo real.
                    try:
                        conn.execute("DELETE FROM supplier_fts")
                        conn.execute(
                            """
                            INSERT INTO supplier_fts(supplier_key, name, email, category, items)
                            SELECT
                                s.supplier_key,
                                COALESCE(s.name_norm, s.name, ''),
                                COALESCE(s.email_norm, s.email, ''),
                                COALESCE(s.raw_category, ''),
                                COALESCE(group_concat(i.name_norm, ' '), '')
                            FROM suppliers s
                            LEFT JOIN supplier_items si ON si.supplier_key = s.supplier_key
                            LEFT JOIN items i ON i.item_id = si.item_id
                            GROUP BY s.supplier_key
                            """
                        )
                    except Exception:
                        # Não deixa uma falha no módulo FTS impedir o cadastro básico.
                        pass

                    fp = self.fingerprint(xlsx_path)
                    self._set_meta(conn, "xlsx_fingerprint", fp)
                    self._set_meta(conn, "last_indexed_at", time.strftime("%Y-%m-%dT%H:%M:%S"))
                    conn.execute("COMMIT")
                p("Catalogo indexado")
                return True, "ok"
            except Exception as e:
                try:
                    with self.connect() as conn:
                        conn.execute("ROLLBACK")
                except Exception:
                    pass
                return False, f"erro_reindex: {e}"

    def _expanded_terms(self, query: str) -> list[str]:
        q = normalize_text(query or "")
        if not q:
            return [""]
        out = {q}
        with self.connect() as conn:
            # Strict expansion:
            # - exact alias -> base
            # - exact base -> aliases
            # - controlled prefix expansion (for short partial terms like "chap")
            rows_alias_exact = conn.execute(
                "SELECT term_base FROM synonyms WHERE alias_norm = ? LIMIT 20",
                (q,),
            ).fetchall()
            rows_base_exact = conn.execute(
                "SELECT alias_norm FROM synonyms WHERE term_base = ? LIMIT 30",
                (q,),
            ).fetchall()
            rows_alias_prefix = conn.execute(
                "SELECT term_base FROM synonyms WHERE alias_norm LIKE ? LIMIT 20",
                (f"{q}%",),
            ).fetchall()
            rows_base_prefix = conn.execute(
                "SELECT term_base FROM synonyms WHERE term_base LIKE ? LIMIT 20",
                (f"{q}%",),
            ).fetchall()

            for rows in (rows_alias_exact, rows_base_exact, rows_alias_prefix, rows_base_prefix):
                for r in rows:
                    term = normalize_text(str(r[0] or ""))
                    if term:
                        out.add(term)
        return sorted(out, key=len, reverse=True)

    def _words(self, text: str) -> list[str]:
        return WORD_RE.findall(normalize_text(text or ""))

    def _has_prefix(self, words: list[str], token: str) -> bool:
        if not token:
            return False
        return any(w.startswith(token) for w in words)

    def _strict_variant_score(
        self,
        *,
        variant: str,
        name_norm: str,
        email_norm: str,
        city_norm: str,
        uf_norm: str,
        category_norm: str,
        item_norm: str,
    ) -> int:
        tokens = self._words(variant)
        if not tokens:
            return 1

        name_words = self._words(name_norm)
        # E-mail split in words to enable prefix matching on local/domain parts.
        email_words = self._words(email_norm.replace("@", " ").replace(".", " "))
        city_words = self._words(f"{city_norm} {uf_norm}")
        category_words = self._words(category_norm)
        item_words = self._words(item_norm)

        score = 0
        for tok in tokens:
            matched = False
            if self._has_prefix(item_words, tok):
                score += 40
                matched = True
            elif self._has_prefix(category_words, tok):
                score += 30
                matched = True
            elif self._has_prefix(name_words, tok):
                score += 20
                matched = True
            elif self._has_prefix(city_words, tok):
                score += 10
                matched = True
            elif self._has_prefix(email_words, tok):
                score += 8
                matched = True

            # Strict mode: all tokens must match at least one indexed field.
            if not matched:
                return 0

        phrase = normalize_text(variant)
        if phrase and phrase in f"{item_norm} {category_norm}":
            score += 14
        if phrase and phrase in name_norm:
            score += 8
        return score

    def _broad_variant_score(
        self,
        *,
        variant: str,
        name_norm: str,
        email_norm: str,
        city_norm: str,
        uf_norm: str,
        category_norm: str,
        item_norm: str,
    ) -> int:
        tokens = self._words(variant)
        if not tokens:
            return 1
        blob = " ".join([name_norm, email_norm, city_norm, uf_norm, category_norm, item_norm])
        if any(tok not in blob for tok in tokens):
            return 0
        score = 0
        for tok in tokens:
            if tok in item_norm:
                score += 26
            elif tok in category_norm:
                score += 18
            elif tok in name_norm:
                score += 12
            elif tok in city_norm or tok in uf_norm:
                score += 6
            elif tok in email_norm:
                score += 4
        return score

    def _fts_match_expr(self, query: str) -> str:
        tokens = [t for t in self._words(query) if len(t) >= 2]
        if not tokens:
            return ""
        # Prefix search para digitação parcial: chapa -> chapa*.
        return " ".join(f"{token}*" for token in tokens[:6])

    def _fts_candidate_keys(self, conn: sqlite3.Connection, query: str, *, limit: int = 1500) -> set[str]:
        expr = self._fts_match_expr(query)
        if not expr:
            return set()
        try:
            rows = conn.execute(
                "SELECT supplier_key FROM supplier_fts WHERE supplier_fts MATCH ? LIMIT ?",
                (expr, max(1, int(limit))),
            ).fetchall()
            return {str(r[0] or "") for r in rows if str(r[0] or "")}
        except Exception:
            return set()

    def list_categories(self) -> list[str]:
        self.init_schema()
        with self.connect() as conn:
            rows = conn.execute(
                "SELECT DISTINCT category FROM items WHERE category IS NOT NULL AND trim(category)<>'' ORDER BY category"
            ).fetchall()
        return [str(r[0]) for r in rows]

    def query_suppliers(
        self,
        query: str,
        *,
        category: str = "",
        limit: int = 100,
        offset: int = 0,
        broad_mode: bool = False,
    ) -> list[CatalogSupplierRow]:
        t0 = time.perf_counter()
        self.init_schema()
        q_variants = self._expanded_terms(query)
        category_norm = normalize_text(category or "")
        rows_out: dict[str, CatalogSupplierRow] = {}

        with self.connect() as conn:
            candidate_keys = self._fts_candidate_keys(conn, query) if normalize_text(query or "") else set()
            sql = (
                """
                SELECT
                    s.supplier_key,
                    s.supplier_id,
                    s.name,
                    s.name_norm,
                    s.email,
                    s.email_norm,
                    s.city,
                    s.uf,
                    COALESCE(s.raw_category, '') AS raw_category,
                    COALESCE(group_concat(i.name, ' '), '') AS item_blob,
                    COALESCE(group_concat(i.category, ' '), '') AS item_categories
                FROM suppliers s
                LEFT JOIN supplier_items si ON si.supplier_key = s.supplier_key
                LEFT JOIN items i ON i.item_id = si.item_id
                """
            )
            params: list = []
            where_parts: list[str] = []
            if candidate_keys:
                placeholders = ",".join("?" for _ in candidate_keys)
                where_parts.append(f"s.supplier_key IN ({placeholders})")
                params.extend(sorted(candidate_keys))
            if category_norm:
                where_parts.append(
                    """
                    (
                        lower(s.raw_category) LIKE ?
                        OR EXISTS (
                            SELECT 1
                            FROM supplier_items sx
                            JOIN items ix ON ix.item_id = sx.item_id
                            WHERE sx.supplier_key = s.supplier_key
                              AND lower(ix.category) LIKE ?
                        )
                    )
                    """
                )
                params.extend([f"%{category_norm}%", f"%{category_norm}%"])
            if where_parts:
                sql += " WHERE " + " AND ".join(where_parts)
            sql += " GROUP BY s.supplier_key"

            all_rows = conn.execute(sql, params).fetchall()

            for r in all_rows:
                key = str(r["supplier_key"] or "")
                if not key:
                    continue

                name = str(r["name"] or "")
                name_norm = normalize_text(str(r["name_norm"] or name))
                email = str(r["email"] or "")
                email_norm = normalize_text(str(r["email_norm"] or email))
                city = str(r["city"] or "")
                city_norm = normalize_text(city)
                uf = str(r["uf"] or "")
                uf_norm = normalize_text(uf)
                raw_category = normalize_text(str(r["raw_category"] or ""))
                item_blob = normalize_text(str(r["item_blob"] or ""))
                item_categories = normalize_text(str(r["item_categories"] or ""))
                category_blob = normalize_text(f"{raw_category} {item_categories}")

                best_score = 0
                for variant in q_variants[:6]:
                    if broad_mode:
                        score = self._broad_variant_score(
                            variant=variant,
                            name_norm=name_norm,
                            email_norm=email_norm,
                            city_norm=city_norm,
                            uf_norm=uf_norm,
                            category_norm=category_blob,
                            item_norm=item_blob,
                        )
                    else:
                        score = self._strict_variant_score(
                            variant=variant,
                            name_norm=name_norm,
                            email_norm=email_norm,
                            city_norm=city_norm,
                            uf_norm=uf_norm,
                            category_norm=category_blob,
                            item_norm=item_blob,
                        )
                    best_score = max(best_score, score)

                # Empty query should list all suppliers ordered by quality signals.
                if normalize_text(query or "") == "":
                    best_score = max(best_score, 1)
                    if email and "@" in email:
                        best_score += 2
                    if item_blob:
                        best_score += 2

                if best_score <= 0:
                    continue

                rows_out[key] = CatalogSupplierRow(
                    supplier_key=key,
                    supplier_id=str(r["supplier_id"] or ""),
                    name=name,
                    email=email,
                    city=city,
                    uf=uf,
                    category=str(r["raw_category"] or ""),
                    base_score=int(best_score),
                )

        rows = list(rows_out.values())
        rows.sort(key=lambda x: (-x.base_score, normalize_text(x.name)))
        offset_i = max(0, int(offset))
        limit_i = max(1, int(limit))
        out = rows[offset_i : offset_i + limit_i]
        record_timing("catalog.query_suppliers_ms", (time.perf_counter() - t0) * 1000.0)
        return out

    def count_suppliers(self, query: str, *, category: str = "", broad_mode: bool = False) -> int:
        # Keep count aligned with strict/broad matching logic.
        rows = self.query_suppliers(query, category=category, limit=1000000, offset=0, broad_mode=broad_mode)
        return len(rows)
