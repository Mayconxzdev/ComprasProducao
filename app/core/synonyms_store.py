from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Set

from openpyxl import load_workbook

from .config import AppConfig, ensure_app_data_dir
from .utils_text import normalize_text


@dataclass
class SynonymMap:
    alias_to_base: Dict[str, str]


class SynonymStore:
    def __init__(self, config: AppConfig):
        self.config = config
        self._cache = SynonymMap(alias_to_base={})
        self._loaded = False

    @property
    def json_path(self) -> Path:
        return ensure_app_data_dir() / "synonyms_local.json"

    def _xlsx_path(self) -> Path | None:
        if self.config.xlsx_sources:
            return Path(self.config.xlsx_sources[0])
        p = (self.config.nas_master_path or "").strip()
        if p.lower().endswith((".xlsx", ".xlsm")):
            return Path(p)
        return None

    def _load_json(self) -> Dict[str, str]:
        if not self.json_path.exists():
            return {}
        try:
            data = json.loads(self.json_path.read_text(encoding="utf-8"))
            out: Dict[str, str] = {}
            if isinstance(data, list):
                for row in data:
                    if not isinstance(row, dict):
                        continue
                    base = normalize_text(str(row.get("termo_base") or ""))
                    alias = normalize_text(str(row.get("alias") or ""))
                    if base and alias:
                        out[alias] = base
            return out
        except Exception:
            return {}

    def _load_xlsx_synonyms(self, xlsx_path: Path) -> Dict[str, str]:
        if not xlsx_path.exists():
            return {}
        wb = load_workbook(str(xlsx_path), data_only=True, read_only=True)
        try:
            out: Dict[str, str] = {}
            # Optional user-friendly sheet.
            if "SINONIMOS" in wb.sheetnames:
                ws = wb["SINONIMOS"]
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    headers = [normalize_text(str(h or "")) for h in rows[0]]
                    try:
                        base_i = headers.index("termo_base")
                        alias_i = headers.index("alias")
                    except ValueError:
                        base_i = 0
                        alias_i = 1
                    for row in rows[1:]:
                        if not row:
                            continue
                        base = normalize_text(str(row[base_i] or "")) if base_i < len(row) else ""
                        alias = normalize_text(str(row[alias_i] or "")) if alias_i < len(row) else ""
                        if base and alias:
                            out[alias] = base
            # Compatibility with professional TERMOS sheet.
            if "TERMOS" in wb.sheetnames:
                ws = wb["TERMOS"]
                for row in ws.iter_rows(values_only=True):
                    if not row or len(row) < 2:
                        continue
                    tipo = normalize_text(str(row[0] or ""))
                    termo = normalize_text(str(row[1] or ""))
                    if tipo == "sinonimo" and termo:
                        # TERMOS only maps term->item_id. Keep identity base for query expansion.
                        out.setdefault(termo, termo)
            return out
        finally:
            wb.close()

    def load(self, force: bool = False) -> SynonymMap:
        if self._loaded and not force:
            return self._cache
        merged = self._load_json()
        xlsx = self._xlsx_path()
        if xlsx is not None:
            try:
                merged.update(self._load_xlsx_synonyms(xlsx))
            except Exception:
                pass
        self._cache = SynonymMap(alias_to_base=merged)
        self._loaded = True
        return self._cache

    def expand_query(self, query: str) -> Set[str]:
        q = normalize_text(str(query or ""))
        if not q:
            return {""}
        m = self.load().alias_to_base
        variants = {q}
        replaced = q
        for alias, base in sorted(m.items(), key=lambda x: len(x[0]), reverse=True):
            if alias and alias in replaced:
                replaced = replaced.replace(alias, base)
        variants.add(" ".join(replaced.split()))
        return {v for v in variants if v}
